# -*- coding: utf-8 -*-
"""Fill Achievement.xlsx icon/frame columns with Resources-compatible asset paths."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
ACHIEVEMENT_XLSX = ROOT / "DataTables" / "Datas" / "Achievement.xlsx"

FRAME_PATH = "Assets/Res/Resources/Textures/UI/Panel/Recruit/frame.png"

CATEGORY_ICON_PATHS = {
    "BusinessMilestone": "Assets/Res/Resources/Textures/UI/TechTree/diandan1.png",
    "StaffGrowth": "Assets/Res/Resources/Textures/UI/TechTree/shouzhang1.png",
    "Expansion": "Assets/Res/Resources/Textures/UI/TechTree/guitai1.png",
    "ContentUnlock": "Assets/Res/Resources/Textures/UI/TechTree/shangcai1.png",
    "VipGameplay": "Assets/Res/Resources/Textures/UI/TechTree/guike1.png",
    "Challenge": "Assets/Res/Resources/Textures/UI/TechTree/lanke1.png",
}

DEFAULT_ICON_PATH = CATEGORY_ICON_PATHS["BusinessMilestone"]


def main() -> None:
    wb = load_workbook(ACHIEVEMENT_XLSX)
    ws = wb.active

    header = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    icon_col = header.index("icon") + 1
    frame_col = header.index("frame") + 1
    category_col = header.index("category") + 1

    updated = 0
    for row in range(5, ws.max_row + 1):
        achievement_id = ws.cell(row, 2).value
        if achievement_id in (None, ""):
            continue

        category = ws.cell(row, category_col).value
        icon_path = CATEGORY_ICON_PATHS.get(str(category), DEFAULT_ICON_PATH)
        if ws.cell(row, icon_col).value != icon_path:
            ws.cell(row, icon_col).value = icon_path
            updated += 1
        if ws.cell(row, frame_col).value != FRAME_PATH:
            ws.cell(row, frame_col).value = FRAME_PATH
            updated += 1

    try:
        wb.save(ACHIEVEMENT_XLSX)
        print(f"patched {ACHIEVEMENT_XLSX} ({updated} cells updated)")
    except PermissionError:
        pending = ACHIEVEMENT_XLSX.with_name("Achievement.pending.xlsx")
        wb.save(pending)
        print(f"WARNING: {ACHIEVEMENT_XLSX} is locked. Wrote {pending} — close Excel and rerun.")


if __name__ == "__main__":
    main()
