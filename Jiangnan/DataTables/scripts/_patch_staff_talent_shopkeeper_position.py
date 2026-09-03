# -*- coding: utf-8 -*-
"""Set StaffTalent 1301-1312 position to Shopkeeper (remove Management usage)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
TALENT_PATH = ROOT / "DataTables" / "Datas" / "StaffTalent.xlsx"
ENUMS_PATH = ROOT / "DataTables" / "Datas" / "__enums__.xlsx"


def patch_talent_positions() -> None:
    wb = load_workbook(TALENT_PATH)
    ws = wb.active
    changed = 0
    for row in range(1, ws.max_row + 1):
        talent_id = ws.cell(row, 2).value
        if not isinstance(talent_id, int) or talent_id < 1301 or talent_id > 1312:
            continue
        if ws.cell(row, 5).value != "Shopkeeper":
            ws.cell(row, 5).value = "Shopkeeper"
            changed += 1
    wb.save(TALENT_PATH)
    print(f"updated {changed} StaffTalent rows to Shopkeeper in {TALENT_PATH}")


def remove_staff_position_management() -> None:
    wb = load_workbook(ENUMS_PATH)
    ws = wb.active
    remove_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 8).value == "Management" and ws.cell(row, 10).value == 4:
            # Only remove StaffPosition.Management (value 4), not StaffTalentType.Management (24)
            prev_is_staff_position = False
            for back in range(row - 1, max(0, row - 5), -1):
                if ws.cell(back, 2).value == "StaffPosition":
                    prev_is_staff_position = True
                    break
                if ws.cell(back, 2).value not in (None, ""):
                    break
            if prev_is_staff_position or (
                row > 1 and ws.cell(row - 1, 8).value in ("Waiter", "Chef", "Shopkeeper")
            ):
                remove_row = row
                break
    if remove_row is None:
        print("StaffPosition.Management row not found (may already be removed)")
        return
    ws.delete_rows(remove_row, 1)
    wb.save(ENUMS_PATH)
    print(f"removed StaffPosition.Management at row {remove_row}")


def main() -> None:
    patch_talent_positions()
    remove_staff_position_management()


if __name__ == "__main__":
    main()
