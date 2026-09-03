# -*- coding: utf-8 -*-
"""Create VipGuestDemandHint.xlsx and register Luban schema."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATAS = ROOT / "DataTables" / "Datas"
DEFINES = ROOT / "DataTables" / "Defines"
CONFIG = ROOT / "Assets" / "Res" / "Resources" / "Config"

VIP_HINT_BEAN_XML = """<module name="">
    <bean name="VipGuestDemandHint">
        <var name="id" type="int"/>
        <var name="text" type="string"/>
    </bean>
</module>
"""

HINT_ROWS = [
    (1, "最近馋{0}菜了"),
    (2, "突然很想吃{0}菜"),
    (3, "近期格外想吃{0}菜"),
    (4, "好久没吃，想吃{0}菜"),
    (5, "心心念念想吃{0}菜"),
    (6, "这会儿特别想吃{0}菜"),
    (7, "总想安排上一盘{0}菜"),
    (8, "最近特别馋这道{0}菜"),
    (9, "迫不及待想吃{0}菜"),
    (10, "闲来无事，想吃{0}菜"),
    (11, "格外爱吃{0}菜"),
    (12, "我超爱吃{0}菜"),
    (13, "一直很爱吃{0}菜"),
    (14, "下饭首选就是{0}菜"),
    (15, "很难拒绝一盘{0}菜"),
    (16, "日常最爱吃{0}菜"),
    (17, "每次吃饭都想吃{0}菜"),
    (18, "妥妥爱吃{0}菜"),
    (19, "百吃不厌的{0}菜"),
    (20, "口味偏爱{0}菜"),
    (21, "{0}菜永远是心头好"),
    (22, "{0}菜当属我的心头最爱"),
    (23, "要说最爱，还得是{0}菜"),
    (24, "{0}菜算得上我的本命菜肴"),
    (25, "没有什么能替代{0}菜"),
    (26, "{0}菜是餐桌上的首选"),
    (27, "论下饭，最爱{0}菜"),
    (28, "{0}菜始终百吃不腻"),
    (29, "心中首选菜肴便是{0}菜"),
    (30, "{0}菜是无可替代的美味"),
    (31, "最近想吃一盘热气腾腾的{0}菜"),
    (32, "我爱吃香气十足的{0}菜"),
    (33, "{0}菜是我的最爱，怎么吃都不腻"),
    (34, "隔一段时间就想吃{0}菜"),
    (35, "一口就满足，最爱{0}菜"),
]


def ensure_hint_bean_xml() -> None:
    """Luban bean 定义；当 __beans__.xlsx 被 Excel 占用时仍可通过 Defines 加载。"""
    path = DEFINES / "vip_guest_demand_hint.xml"
    if path.exists() and "VipGuestDemandHint" in path.read_text(encoding="utf-8"):
        print(f"VipGuestDemandHint xml already in {path}")
        return

    path.write_text(VIP_HINT_BEAN_XML, encoding="utf-8")
    print(f"saved {path}")


def ensure_hint_bean() -> None:
    path = DATAS / "__beans__.xlsx"
    try:
        wb = load_workbook(path)
    except PermissionError as error:
        print(f"skip bean patch (locked): {path} ({error})")
        return

    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "VipGuestDemandHint":
            print(f"VipGuestDemandHint bean already in {path}")
            return

    fields = [
        ("id", "int", "模板Id"),
        ("text", "string", "提示语模板"),
    ]
    start = ws.max_row + 1
    for index, (fname, ftype, comment) in enumerate(fields):
        row = start + index
        ws.cell(row, 2).value = "VipGuestDemandHint" if index == 0 else None
        if index == 0:
            ws.cell(row, 7).value = "贵客猜菜提示语"
        ws.cell(row, 10).value = fname
        ws.cell(row, 12).value = ftype
        ws.cell(row, 14).value = comment
    try:
        wb.save(path)
        print(f"patched beans {path}")
    except PermissionError as error:
        print(f"skip bean save (locked): {path} ({error})")


def ensure_hint_table() -> None:
    path = DATAS / "__tables__.xlsx"
    try:
        wb = load_workbook(path)
    except PermissionError as error:
        print(f"skip table patch (locked): {path} ({error})")
        return

    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "TbVipGuestDemandHint":
            print(f"TbVipGuestDemandHint already in {path}")
            return

    row = ws.max_row + 1
    ws.cell(row, 2).value = "TbVipGuestDemandHint"
    ws.cell(row, 3).value = "VipGuestDemandHint"
    ws.cell(row, 5).value = "VipGuestDemandHint.xlsx"
    ws.cell(row, 6).value = "id"
    ws.cell(row, 7).value = "map"
    ws.cell(row, 9).value = "贵客猜菜提示语"
    try:
        wb.save(path)
        print(f"patched tables {path}")
    except PermissionError as error:
        print(f"skip table save (locked): {path} ({error})")


def write_hint_xlsx() -> None:
    path = DATAS / "VipGuestDemandHint.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        ("##var", "id", "text"),
        ("##type", "int", "string"),
        ("##group", "c", "c"),
        ("##", "模板Id", "提示语"),
    ]
    for row_index, row in enumerate(headers, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value

    for index, (hint_id, text) in enumerate(HINT_ROWS):
        excel_row = 5 + index
        ws.cell(excel_row, 1).value = None
        ws.cell(excel_row, 2).value = hint_id
        ws.cell(excel_row, 3).value = text
    wb.save(path)
    print(f"saved {path}")


def write_hint_json() -> None:
    rows = [{"id": hint_id, "text": text} for hint_id, text in HINT_ROWS]
    out = CONFIG / "tbvipguestdemandhint.json"
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"saved {out}")


def main() -> None:
    ensure_hint_bean_xml()
    ensure_hint_bean()
    ensure_hint_table()
    write_hint_xlsx()
    write_hint_json()


if __name__ == "__main__":
    main()
