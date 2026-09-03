#!/usr/bin/env python3
"""Create Facility Luban Excel and register into schema (keeps GuideTask/Staff)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATAS = ROOT / "Datas"


def write_rows(ws, rows: list[list]):
    ws.delete_rows(1, ws.max_row or 1)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val


def ensure_enums():
    path = DATAS / "__enums__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "flags", "unique", "group", "comment", "tags", "*items", None, None, None, None],
            ["##var", None, None, None, None, None, None, "name", "alias", "value", "comment", "tags"],
            ["##", "全名(包含模块名可空)", "是否为位标枚举", "枚举项是否唯一", None, None, None, "枚举名", "别名", "值", "备注", None],
            # GuideTaskType
            [None, "GuideTaskType", None, True, None, "引导任务类型", None, "BuyBasicEquipment", "购买基础设施", 1, None, None],
            [None, None, None, None, None, None, None, "BuyTables", "购买桌子", 2, None, None],
            [None, None, None, None, None, None, None, "BuyKitchenEquipment", "购买厨房设施", 3, None, None],
            [None, None, None, None, None, None, None, "HireShopkeeper", "招聘掌柜", 4, None, None],
            [None, None, None, None, None, None, None, "HireChef", "招聘厨师", 5, None, None],
            [None, None, None, None, None, None, None, "HireWaiter", "招聘小二", 6, None, None],
            [None, None, None, None, None, None, None, "Custom", "自定义", 99, None, None],
            # StaffPosition
            [None, "StaffPosition", None, True, None, "员工职位", None, "Shopkeeper", "掌柜", 1, None, None],
            [None, None, None, None, None, None, None, "Chef", "厨师", 2, None, None],
            [None, None, None, None, None, None, None, "Waiter", "小二", 3, None, None],
            # StaffQuality
            [None, "StaffQuality", None, True, None, "员工品质", None, "Common", "普通", 1, None, None],
            [None, None, None, None, None, None, None, "Good", "优良", 2, None, None],
            [None, None, None, None, None, None, None, "Rare", "稀有", 3, None, None],
            [None, None, None, None, None, None, None, "Epic", "史诗", 4, None, None],
            # FacilityType
            [None, "FacilityType", None, True, None, "设施类型", None, "Table", "桌子", 1, None, None],
            [None, None, None, None, None, None, None, "Counter", "掌柜桌", 2, None, None],
            [None, None, None, None, None, None, None, "Stove", "灶台", 3, None, None],
            [None, None, None, None, None, None, None, "Furnace", "炉子", 4, None, None],
            [None, None, None, None, None, None, None, "WineCabinet", "酒柜", 5, None, None],
            [None, None, None, None, None, None, None, "Cabinet", "柜子", 6, None, None],
            [None, None, None, None, None, None, None, "KitchenTable", "厨房桌", 7, None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_beans():
    path = DATAS / "__beans__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "parent", "valueType", "sep", "alias", "comment", "group", "tags", "*fields", None, None, None, None, None, None],
            ["##var", None, None, None, None, None, None, None, None, "name", "alias", "type", "group", "comment", "tags", "variants"],
            ["##", "全名(包含模块名可空)", None, None, "分隔符", None, None, None, None, "字段名", "字段别名", "类型", "分组", "备注", None, "字段变体"],
            # Config
            [None, "Config", None, None, None, None, None, None, None, "configName", None, "string", None, "配置名称", None, None],
            [None, None, None, None, None, None, None, None, None, "value", None, "int", None, "数值", None, None],
            # GuideTask
            [None, "GuideTask", None, None, None, None, "引导小目标任务", None, None, "id", None, "int", None, "任务Id", None, None],
            [None, None, None, None, None, None, None, None, None, "desc", None, "string", None, "描述文字", None, None],
            [None, None, None, None, None, None, None, None, None, "taskType", None, "GuideTaskType", None, "任务类型", None, None],
            [None, None, None, None, None, None, None, None, None, "param", None, "(list#sep=,),int", None, "关联参数列表(按taskType解释)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "解锁前置任务Id列表", None, None],
            # Staff
            [None, "Staff", None, None, None, None, "员工属性", None, None, "id", None, "int", None, "员工Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "姓名", None, None],
            [None, None, None, None, None, None, None, None, None, "position", None, "StaffPosition", None, "职位", None, None],
            [None, None, None, None, None, None, None, None, None, "moveSpeed", None, "float", None, "移动速度", None, None],
            [None, None, None, None, None, None, None, None, None, "canOrder", None, "bool", None, "会点单", None, None],
            [None, None, None, None, None, None, None, None, None, "canServe", None, "bool", None, "会上菜", None, None],
            [None, None, None, None, None, None, None, None, None, "canCheckout", None, "bool", None, "会收账", None, None],
            [None, None, None, None, None, None, None, None, None, "serviceAttitude", None, "int", None, "服务态度指数", None, None],
            [None, None, None, None, None, None, None, None, None, "personality", None, "int", None, "性格指数", None, None],
            [None, None, None, None, None, None, None, None, None, "quality", None, "StaffQuality", None, "品质", None, None],
            [None, None, None, None, None, None, None, None, None, "salary", None, "int", None, "工资", None, None],
            [None, None, None, None, None, None, None, None, None, "visual", None, "string", None, "任务形象/预制体Key", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
            # Facility
            [None, "Facility", None, None, None, None, "酒楼设施", None, None, "id", None, "int", None, "设施Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "显示名", None, None],
            [None, None, None, None, None, None, None, None, None, "facilityType", None, "FacilityType", None, "设施类型", None, None],
            [None, None, None, None, None, None, None, None, None, "bindNode", None, "string", None, "场景绑定节点名", None, None],
            [None, None, None, None, None, None, None, None, None, "guideKey", None, "string", None, "引导逻辑Key", None, None],
            [None, None, None, None, None, None, None, None, None, "equipmentId", None, "int", None, "关联SO_EquipmentId(无则0)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "解锁前置设施Id列表", None, None],
            [None, None, None, None, None, None, None, None, None, "unlockIncome", None, "int", None, "解锁所需累计营收(0忽略)", None, None],
            [None, None, None, None, None, None, None, None, None, "cost", None, "int", None, "解锁花费", None, None],
            [None, None, None, None, None, None, None, None, None, "upgradeGroup", None, "(list#sep=,),int", None, "升级用设备Id组(SO_Equipment)", None, None],
            [None, None, None, None, None, None, None, None, None, "sortOrder", None, "int", None, "同类型排序", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_tables():
    path = DATAS / "__tables__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "value_type", "read_schema_from_file", "input", "index", "mode", "group", "comment", "tags", "output"],
            ["##", "全名(包含模块名可空)", "记录类型", "从excel读取定义", "文件列表", "主键字段", "模式", "分组", "备注", None, "输出文件名"],
            ["##", None, None, None, None, None, None, None, None, None, None],
            [None, "TbConfig", "Config", None, "Config.xlsx", None, "list", None, "全局KV配置", None, None],
            [None, "TbGuideTask", "GuideTask", None, "GuideTask.xlsx", "id", "map", None, "引导任务/小目标", None, None],
            [None, "TbStaff", "Staff", None, "Staff.xlsx", "id", "map", None, "员工属性", None, None],
            [None, "TbFacility", "Facility", None, "Facility.xlsx", "id", "map", None, "酒楼设施", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def create_facility_xlsx():
    path = DATAS / "Facility.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # unlock: 前置设施Id；unlockIncome: 累计营收门槛；upgradeGroup: SO_Equipment Id 列表
    header = [
        ["##var", "id", "name", "facilityType", "bindNode", "guideKey", "equipmentId", "unlock", "unlockIncome", "cost", "upgradeGroup", "sortOrder", "remark"],
        ["##type", "int", "string", "FacilityType", "string", "string", "int", "(list#sep=,),int", "int", "int", "(list#sep=,),int", "int", "string"],
        ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
        ["##", "设施Id", "显示名", "类型", "绑定节点", "引导Key", "设备Id", "前置设施", "营收门槛", "花费", "升级设备组", "排序", "备注"],
    ]
    # 桌子 1-6：引导前 2 张无营收门槛；3-6 按累计营收逐渐解锁（可改表）
    rows = header + [
        [None, 1, "桌子1", "Table", "TableArea_1", "table_1", 2, "", 0, 900, "2", 1, "引导期首桌"],
        [None, 2, "桌子2", "Table", "TableArea_2", "table_2", 2, "1", 0, 900, "2", 2, "引导期第二桌"],
        [None, 3, "桌子3", "Table", "TableArea_3", "table_3", 2, "2", 800, 900, "2", 3, "开业后按营收解锁"],
        [None, 4, "桌子4", "Table", "TableArea_4", "table_4", 2, "3", 2000, 900, "2", 4, None],
        [None, 5, "桌子5", "Table", "TableArea_5", "table_5", 2, "4", 4000, 900, "2", 5, None],
        [None, 6, "桌子6", "Table", "TableArea_6", "table_6", 2, "5", 7000, 900, "2", 6, None],
        # 基础设施 / 厨房
        [None, 10, "掌柜桌", "Counter", "柜台", "counter", 0, "", 0, 0, "0", 1, "花费0表示走SO_Equipment.lv1"],
        [None, 11, "柜子", "Cabinet", "柜子", "cabinet", 0, "10", 0, 0, "", 2, None],
        [None, 12, "酒柜", "WineCabinet", "酒柜", "wine_cabinet", 0, "11", 0, 0, "", 3, None],
        [None, 20, "灶台", "Stove", "BigStove", "stove", 3, "", 0, 0, "3", 1, None],
        [None, 21, "炉子", "Furnace", "SmallStove", "furnace", 0, "20", 0, 0, "", 2, None],
        [None, 22, "厨房桌子1", "KitchenTable", "厨房桌子1", "kitchen_table_1", 0, "21", 0, 0, "", 3, None],
        [None, 23, "厨房桌子2", "KitchenTable", "厨房桌子2", "kitchen_table_2", 0, "22", 0, 0, "", 4, None],
    ]
    write_rows(ws, rows)
    try:
        wb.save(path)
    except PermissionError:
        print(f"skip {path} (locked)")
        return
    print(f"created {path}")


def main():
    DATAS.mkdir(parents=True, exist_ok=True)
    ensure_enums()
    ensure_beans()
    ensure_tables()
    create_facility_xlsx()


if __name__ == "__main__":
    main()
