# -*- coding: utf-8 -*-
"""Expand Staff hire pool and sync tbstaff.json."""
from openpyxl import load_workbook
from pathlib import Path
import json

ROOT = Path(__file__).resolve().parents[2]

# id, name, position, moveSpeed, canOrder, canServe, canCheckout,
# attitude, personality, quality, salary, visual, remark
ROWS = [
    (1, "阿福", "Shopkeeper", 2.2, True, False, True, 80, 60, "Good", 120, "WaiterF1", "默认掌柜·前台接待收账"),
    (4, "老李", "Chef", 1.8, False, False, False, 70, 55, "Good", 150, "Chef3", "厨师·稳重手艺好"),
    (8, "阿旺", "Chef", 1.6, False, False, False, 55, 45, "Common", 90, "Chef3", "厨师·便宜勤快慢"),
    (9, "阿强", "Chef", 1.7, False, False, False, 60, 50, "Common", 110, "Chef3", "厨师·中等均衡"),
    (10, "铁牛", "Chef", 2.0, False, False, False, 48, 40, "Common", 95, "Chef3", "厨师·手脚快态度一般"),
    (11, "阿贵", "Chef", 1.5, False, False, False, 78, 62, "Common", 125, "Chef3", "厨师·慢工出细活"),
    (12, "小厨", "Chef", 1.9, False, False, False, 58, 70, "Common", 85, "Chef3", "厨师·年轻活泼低薪"),
    (13, "周师傅", "Chef", 1.75, False, False, False, 82, 58, "Good", 180, "Chef3", "厨师·优良老手"),
    (14, "陈掌勺", "Chef", 1.65, False, False, False, 65, 48, "Common", 100, "Chef3", "厨师·性价比"),
    (15, "阿福厨", "Chef", 1.85, False, False, False, 88, 72, "Rare", 220, "Chef3", "厨师·稀有高薪"),
    (5, "小荷", "Waiter", 2.5, True, False, False, 75, 65, "Common", 100, "WaiterF1_1", "小二·均衡默认"),
    (6, "阿翠", "Waiter", 2.2, True, False, False, 55, 50, "Common", 70, "WaiterF1_1", "小二·低薪入门"),
    (7, "阿柳", "Waiter", 2.3, True, False, False, 60, 55, "Common", 80, "WaiterF1", "小二·中低阶"),
    (16, "小桃", "Waiter", 2.8, True, False, False, 52, 68, "Common", 95, "WaiterF1_1", "小二·跑得快"),
    (17, "阿梅", "Waiter", 2.1, True, False, False, 80, 72, "Common", 115, "WaiterF1", "小二·态度好性格佳"),
    (18, "阿兰", "Waiter", 2.0, True, False, False, 70, 45, "Common", 90, "WaiterF1_1", "小二·稳重慢一点"),
    (19, "小芸", "Waiter", 2.6, True, False, False, 85, 70, "Good", 140, "WaiterF1", "小二·优良服务"),
    (20, "阿珍", "Waiter", 2.4, True, False, False, 50, 42, "Common", 65, "WaiterF1_1", "小二·最便宜"),
    (21, "小月", "Waiter", 2.55, True, False, False, 68, 60, "Common", 105, "WaiterF1", "小二·中上均衡"),
    (22, "阿霞", "Waiter", 2.35, True, False, False, 90, 78, "Rare", 200, "WaiterF1_1", "小二·稀有高薪"),
]

POS = {"Shopkeeper": 1, "Chef": 2, "Waiter": 3}
QUAL = {"Common": 1, "Good": 2, "Rare": 3, "Epic": 4}


def main():
    staff_xlsx = ROOT / "DataTables" / "Datas" / "Staff.xlsx"

    wb = load_workbook(staff_xlsx)
    ws = wb.active
    if ws.max_row > 4:
        ws.delete_rows(5, ws.max_row - 4)

    for i, row in enumerate(ROWS):
        row_idx = 5 + i
        ws.cell(row_idx, 1, None)
        for col, value in enumerate(row, start=2):
            ws.cell(row_idx, col, value)

    wb.save(staff_xlsx)
    print(f"xlsx saved {staff_xlsx} rows={len(ROWS)}")

    json_rows = []
    for row in ROWS:
        json_rows.append(
            {
                "id": row[0],
                "name": row[1],
                "position": POS[row[2]],
                "moveSpeed": row[3],
                "canOrder": row[4],
                "canServe": row[5],
                "canCheckout": row[6],
                "serviceAttitude": row[7],
                "personality": row[8],
                "quality": QUAL[row[9]],
                "salary": row[10],
                "visual": row[11],
                "remark": row[12],
            }
        )

    out = ROOT / "Assets" / "Res" / "Resources" / "Config" / "tbstaff.json"
    out.write_text(json.dumps(json_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"json saved {out} count={len(json_rows)}")


if __name__ == "__main__":
    main()
