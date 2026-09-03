# -*- coding: utf-8 -*-
"""Single source of truth: 8-tech tree + staff level skills → Excel + JSON."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATAS = ROOT / "DataTables" / "Datas"
CONFIG = ROOT / "Assets" / "Res" / "Resources" / "Config"

POS = {"Shopkeeper": 1, "Chef": 2, "Waiter": 3}

TECH_TYPE = {
    "ExtraWaiterCap": 1,
    "ExtraChefCap": 2,
    "ExtraShopkeeperCap": 3,
    "QueueCap": 4,
    "BusinessHoursBonus": 5,
    "CustomerRefreshMul": 6,
    "PriceProfitBonus": 7,
    "InspireBonus": 8,
    "ResearchSpeedMul": 9,
    "TipGlobalBonus": 10,
    "UnlockStaffLevel": 11,
    "CustomerRefreshSecondsBonus": 12,
    "BusinessHoursMul": 13,
    "EnableWaiterAutoUpgrade": 14,
    "EnableChefAutoUpgrade": 15,
    "EnableVipCustomer": 16,
    "Custom": 99,
}

# id, position, level, name, move, cook, order, serve, checkout, clean,
# tip, stamina, canOrder, canServe, canCheckout, attitude+, personality+, upgradeCost, remark
STAFF_LEVEL_ROWS = [
    (101, "Waiter", 1, "会点单", 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0, 1.0, True, False, False, 0, 0, 0, "入门：仅会点单"),
    (102, "Waiter", 2, "快点单", 1.05, 1.0, 0.95, 1.0, 1.0, 0.95, 0, 1.0, True, False, False, 3, 0, 0, "加快点单速度"),
    (103, "Waiter", 3, "会上菜", 1.1, 1.0, 0.9, 0.95, 1.0, 0.9, 2, 0.95, True, True, False, 6, 2, 0, "学会上菜"),
    (104, "Waiter", 4, "快上菜", 1.15, 1.0, 0.85, 0.9, 0.9, 0.85, 4, 0.95, True, True, False, 8, 4, 0, "加快上菜速度"),
    (105, "Waiter", 5, "会收账", 1.2, 1.0, 0.8, 0.85, 0.85, 0.8, 6, 0.9, True, True, True, 12, 6, 0, "学会收账"),
    (106, "Waiter", 6, "快收账", 1.25, 1.0, 0.75, 0.8, 0.8, 0.75, 8, 0.9, True, True, True, 15, 8, 0, "加快收账速度"),
    (107, "Waiter", 7, "疾行小二", 1.3, 1.0, 0.7, 0.75, 0.75, 0.7, 12, 0.85, True, True, True, 18, 10, 0, "加快移速"),
    (108, "Waiter", 8, "店中翘楚", 1.35, 1.0, 0.65, 0.7, 0.7, 0.65, 15, 0.8, True, True, True, 22, 12, 0, "移速与服务全面提升"),
    (201, "Chef", 1, "见习厨师", 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0, 1.0, False, False, False, 0, 0, 0, "会作基础菜（红烧肉）"),
    (202, "Chef", 2, "掌勺学徒", 1.03, 1.12, 1.0, 1.0, 1.0, 1.0, 0, 0.97, False, False, False, 0, 3, 0, "研发新菜·包子；加快做菜"),
    (203, "Chef", 3, "熟练厨师", 1.06, 1.22, 1.0, 1.0, 1.0, 1.0, 0, 0.94, False, False, False, 2, 5, 0, "研发新菜·鱼；做菜更快"),
    (204, "Chef", 4, "快手厨", 1.1, 1.35, 1.0, 1.0, 1.0, 1.0, 0, 0.9, False, False, False, 4, 8, 0, "研发新菜·面条；高峰不堵灶"),
    (205, "Chef", 5, "金牌大厨", 1.12, 1.48, 1.0, 1.0, 1.0, 1.0, 0, 0.86, False, False, False, 6, 10, 0, "研发新菜·酒；多桌仍稳"),
    (206, "Chef", 6, "镇店厨神", 1.15, 1.6, 1.0, 1.0, 1.0, 1.0, 0, 0.82, False, False, False, 8, 12, 0, "研发新菜·麻婆豆腐；满编制缓冲"),
    (301, "Shopkeeper", 1, "见习掌柜", 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0, 1.0, False, False, True, 0, 0, 0, "默认可收账"),
    (302, "Shopkeeper", 2, "熟练掌柜", 1.05, 1.0, 1.0, 1.0, 0.9, 1.0, 3, 0.95, False, False, True, 5, 5, 0, "收账更快"),
    (303, "Shopkeeper", 3, "金牌掌柜", 1.1, 1.0, 1.0, 1.0, 0.8, 1.0, 6, 0.9, False, False, True, 10, 10, 0, "收账与小费提升"),
]

# id, name, desc, techType, param, unlock, cost, researchServeCustomers, sortOrder, remark
TECH_ROWS = [
    (
        101,
        "小二开智",
        "小二完成服务行为后自动升级：会点单→加快点单→会上菜→加快上菜→会收账→加快收账→移速提升",
        "EnableWaiterAutoUpgrade",
        "",
        "",
        200,
        20,
        101,
        "员工·行为自动升级",
    ),
    (
        102,
        "刷客间隔缩短",
        "刷客间隔缩短4秒：(customerRefreshTime-4)/解锁桌数",
        "CustomerRefreshSecondsBonus",
        "4",
        "101",
        600,
        35,
        102,
        "客流",
    ),
    (
        103,
        "厨师开智",
        "厨师每完成若干次做菜自动升级：研发新菜、加快做菜速度",
        "EnableChefAutoUpgrade",
        "",
        "102",
        300,
        25,
        103,
        "员工·厨师自动升级",
    ),
    (
        104,
        "延时打烊",
        "营业时间翻倍",
        "BusinessHoursMul",
        "2000",
        "103",
        900,
        40,
        104,
        "客流",
    ),
    (
        105,
        "小费",
        "全店结账小费+5%",
        "TipGlobalBonus",
        "5",
        "104",
        700,
        35,
        105,
        "经营",
    ),
    (
        106,
        "涨价",
        "菜品收入永久+5%",
        "PriceProfitBonus",
        "5",
        "105",
        1200,
        45,
        106,
        "经营",
    ),
    (
        107,
        "拉客贵客",
        "解锁贵客判定玩法（规则待配置）",
        "EnableVipCustomer",
        "",
        "106",
        1500,
        55,
        107,
        "客流·占位",
    ),
    (
        108,
        "开启二楼",
        "暂未开放",
        "Custom",
        "",
        "",
        0,
        0,
        108,
        "经营·锁定",
    ),
]

STAFF_LEVEL_HEADERS = [
    (
        "##var",
        "id",
        "position",
        "level",
        "name",
        "moveSpeedMul",
        "cookSpeedMul",
        "orderTimeMul",
        "serveTimeMul",
        "checkoutTimeMul",
        "cleanTimeMul",
        "tipBonusPercent",
        "staminaDrainMul",
        "canOrder",
        "canServe",
        "canCheckout",
        "serviceAttitudeBonus",
        "personalityBonus",
        "upgradeCost",
        "remark",
    ),
    (
        "##type",
        "int",
        "StaffPosition",
        "int",
        "string",
        "float",
        "float",
        "float",
        "float",
        "float",
        "float",
        "int",
        "float",
        "bool",
        "bool",
        "bool",
        "int",
        "int",
        "int",
        "string",
    ),
    ("##group",) + ("c",) * 19,
    (
        "##",
        "等级配置Id",
        "所属职位",
        "等级",
        "等级显示名",
        "移速倍率",
        "做菜倍率",
        "点单耗时",
        "上菜耗时",
        "收账耗时",
        "清扫耗时",
        "小费%",
        "体力消耗",
        "点单",
        "上菜",
        "收账",
        "态度+",
        "性格+",
        "升级花费",
        "备注",
    ),
]

TECH_HEADERS = [
    ("##var", "id", "name", "desc", "techType", "param", "unlock", "cost", "researchServeCustomers", "sortOrder", "remark"),
    (
        "##type",
        "int",
        "string",
        "string",
        "TavernTechType",
        "(list#sep=,),int",
        "(list#sep=,),int",
        "int",
        "int",
        "int",
        "string",
    ),
    ("##group",) + ("c",) * 10,
    ("##", "科技Id", "名称", "描述", "效果类型", "参数", "前置科技", "花费", "招待顾客(结算)数量", "排序", "备注"),
]

# name, alias, value, comment — inserted into TavernTechType before Custom=99
EXTRA_TAVERN_TECH_ENUMS = [
    ("UnlockStaffLevel", "解锁员工等级", 11, None),
    ("CustomerRefreshSecondsBonus", "刷客间隔减秒", 12, "param 为减去的秒数"),
    ("BusinessHoursMul", "营业时长千分比倍率", 13, None),
    ("EnableWaiterAutoUpgrade", "解锁小二行为自动升级", 14, None),
    ("EnableChefAutoUpgrade", "解锁厨师做菜自动升级", 15, None),
    ("EnableVipCustomer", "解锁贵客玩法", 16, None),
]


def parse_int_list(text: str) -> list[int]:
    text = (text or "").strip()
    if not text:
        return []
    return [int(p.strip()) for p in text.split(",") if p.strip()]


def _rewrite_data_sheet(path: Path, headers: list[tuple], data_rows: list[tuple]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r, row in enumerate(headers, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val
    for i, row in enumerate(data_rows):
        r = len(headers) + 1 + i
        ws.cell(r, 1).value = None
        for c, val in enumerate(row, start=2):
            ws.cell(r, c).value = val
    wb.save(path)
    print(f"saved {path} rows={len(data_rows)}")


def write_staff_level_excel() -> None:
    _rewrite_data_sheet(DATAS / "StaffLevel.xlsx", STAFF_LEVEL_HEADERS, STAFF_LEVEL_ROWS)


def write_tech_excel() -> None:
    _rewrite_data_sheet(DATAS / "TavernTech.xlsx", TECH_HEADERS, TECH_ROWS)


def write_staff_level_json() -> None:
    rows = []
    for row in STAFF_LEVEL_ROWS:
        rows.append(
            {
                "id": row[0],
                "position": POS[row[1]],
                "level": row[2],
                "name": row[3],
                "moveSpeedMul": row[4],
                "cookSpeedMul": row[5],
                "orderTimeMul": row[6],
                "serveTimeMul": row[7],
                "checkoutTimeMul": row[8],
                "cleanTimeMul": row[9],
                "tipBonusPercent": row[10],
                "staminaDrainMul": row[11],
                "canOrder": row[12],
                "canServe": row[13],
                "canCheckout": row[14],
                "serviceAttitudeBonus": row[15],
                "personalityBonus": row[16],
                "upgradeCost": row[17],
                "remark": row[18],
            }
        )
    out = CONFIG / "tbstafflevel.json"
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"saved {out} count={len(rows)}")


def write_tech_json() -> None:
    rows = []
    for row in TECH_ROWS:
        rows.append(
            {
                "id": row[0],
                "name": row[1],
                "desc": row[2],
                "techType": TECH_TYPE[row[3]],
                "param": parse_int_list(row[4]),
                "unlock": parse_int_list(row[5]),
                "cost": row[6],
                "researchServeCustomers": row[7],
                "sortOrder": row[8],
                "remark": row[9],
            }
        )
    out = CONFIG / "tbtaverntech.json"
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"saved {out} count={len(rows)}")


def patch_tavern_tech_enums() -> None:
    path = DATAS / "__enums__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    existing_values: set[int] = set()
    custom_row = None
    in_tavern = False
    for row in range(1, ws.max_row + 1):
        full_name = ws.cell(row, 2).value
        item_name = ws.cell(row, 8).value
        value = ws.cell(row, 10).value
        if full_name == "TavernTechType":
            in_tavern = True
            continue
        if in_tavern and full_name not in (None, ""):
            break
        if not in_tavern:
            continue
        if item_name and value is not None:
            existing_values.add(int(value))
            if item_name == "Custom":
                custom_row = row
    if custom_row is None:
        raise RuntimeError("TavernTechType Custom row not found in __enums__.xlsx")
    insert_at = custom_row
    added = 0
    for name, alias, value, comment in EXTRA_TAVERN_TECH_ENUMS:
        if value in existing_values:
            continue
        ws.insert_rows(insert_at)
        ws.cell(insert_at, 2).value = None
        ws.cell(insert_at, 8).value = name
        ws.cell(insert_at, 9).value = alias
        ws.cell(insert_at, 10).value = value
        ws.cell(insert_at, 11).value = comment
        insert_at += 1
        custom_row += 1
        added += 1
    wb.save(path)
    print(f"patched {path} added_enum_items={added}")


def main() -> None:
    raise SystemExit(
        "Deprecated: use DataTables/scripts/_apply_staff_tech_atomic_v3.py instead."
    )


if __name__ == "__main__":
    main()
