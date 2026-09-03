#!/usr/bin/env python3
"""Create GuideTask / Staff Luban Excel tables and register schema."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATAS = ROOT / "Datas"


def write_rows(ws, rows: list[list]):
    ws.delete_rows(1, ws.max_row or 1)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val


def ensure_enums():
    path = DATAS / "__enums__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "flags", "unique", "group", "comment", "tags", "*items", None, None, None, None],
            ["##var", None, None, None, None, None, None, "name", "alias", "value", "comment", "tags"],
            [
                "##",
                "全名(包含模块名可空)",
                "是否为位标枚举",
                "枚举项是否唯一",
                None,
                None,
                None,
                "枚举名",
                "别名",
                "值",
                "备注",
                None,
            ],
            # GuideTaskType
            [None, "GuideTaskType", None, True, None, "引导任务类型", None, "BuyBasicEquipment", "购买基础设施", 1, "购买柜台/酒柜等基础设施", None],
            [None, None, None, None, None, None, None, "BuyTables", "购买桌子", 2, "购买开业所需桌子", None],
            [None, None, None, None, None, None, None, "BuyKitchenEquipment", "购买厨房设施", 3, "购买灶台等厨房设施", None],
            [None, None, None, None, None, None, None, "HireShopkeeper", "招聘掌柜", 4, None, None],
            [None, None, None, None, None, None, None, "HireChef", "招聘厨师", 5, None, None],
            [None, None, None, None, None, None, None, "HireWaiter", "招聘小二", 6, None, None],
            [None, None, None, None, None, None, None, "Custom", "自定义", 99, "扩展用", None],
            # StaffPosition
            [None, "StaffPosition", None, True, None, "员工职位", None, "Shopkeeper", "掌柜", 1, None, None],
            [None, None, None, None, None, None, None, "Chef", "厨师", 2, None, None],
            [None, None, None, None, None, None, None, "Waiter", "小二", 3, None, None],
            # StaffQuality
            [None, "StaffQuality", None, True, None, "员工品质", None, "Common", "普通", 1, None, None],
            [None, None, None, None, None, None, None, "Good", "优良", 2, None, None],
            [None, None, None, None, None, None, None, "Rare", "稀有", 3, None, None],
            [None, None, None, None, None, None, None, "Epic", "史诗", 4, None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_beans():
    path = DATAS / "__beans__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "parent", "valueType", "sep", "alias", "comment", "group", "tags", "*fields", None, None, None, None, None, None],
            ["##var", None, None, None, None, None, None, None, None, "name", "alias", "type", "group", "comment", "tags", "variants"],
            ["##", "全名(包含模块名可空)", None, None, "分隔符", None, None, None, None, "字段名", "字段别名", "类型", "分组", "备注", None, "字段变体"],
            # Config (existing)
            [None, "Config", None, None, None, None, None, None, None, "configName", None, "string", None, "配置名称", None, None],
            [None, None, None, None, None, None, None, None, None, "value", None, "int", None, "数值", None, None],
            # GuideTask
            [None, "GuideTask", None, None, None, None, "引导小目标任务", None, None, "id", None, "int", None, "任务Id", None, None],
            [None, None, None, None, None, None, None, None, None, "desc", None, "string", None, "描述文字", None, None],
            [None, None, None, None, None, None, None, None, None, "taskType", None, "GuideTaskType", None, "任务类型", None, None],
            [None, None, None, None, None, None, None, None, None, "param", None, "(list#sep=,),int", None, "关联参数列表(按taskType解释)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "解锁前置任务Id列表", None, None],
            # Staff
            [None, "Staff", None, None, None, None, "员工属性", None, None, "id", None, "int", None, "员工Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "姓名", None, None],
            [None, None, None, None, None, None, None, None, None, "position", None, "StaffPosition", None, "职位", None, None],
            [None, None, None, None, None, None, None, None, None, "moveSpeed", None, "float", None, "移动速度", None, None],
            [None, None, None, None, None, None, None, None, None, "canOrder", None, "bool", None, "会点单", None, None],
            [None, None, None, None, None, None, None, None, None, "canServe", None, "bool", None, "会上菜", None, None],
            [None, None, None, None, None, None, None, None, None, "canCheckout", None, "bool", None, "会收账", None, None],
            [None, None, None, None, None, None, None, None, None, "serviceAttitude", None, "int", None, "服务态度指数", None, None],
            [None, None, None, None, None, None, None, None, None, "personality", None, "int", None, "性格指数", None, None],
            [None, None, None, None, None, None, None, None, None, "quality", None, "StaffQuality", None, "品质", None, None],
            [None, None, None, None, None, None, None, None, None, "salary", None, "int", None, "工资", None, None],
            [None, None, None, None, None, None, None, None, None, "visual", None, "string", None, "任务形象/预制体Key", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_tables():
    path = DATAS / "__tables__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "value_type", "read_schema_from_file", "input", "index", "mode", "group", "comment", "tags", "output"],
            ["##", "全名(包含模块名可空)", "记录类型", "从excel读取定义", "文件列表", "主键字段", "模式", "分组", "备注", None, "输出文件名"],
            [
                "##",
                None,
                None,
                "false时取表中定义,true为从excel表头及数据中读取定义",
                "可以多文件用','分隔",
                "为空的话自动取value_type中第一个字段",
                "取值one|map|list，为空时默认为map",
                "取值c|s|e，可多分组用','分隔，空表示属于所有分组",
                None,
                None,
                "默认为 <module>_<name>.<suffix>",
            ],
            [None, "TbConfig", "Config", None, "Config.xlsx", None, "list", None, "全局KV配置", None, None],
            [None, "TbGuideTask", "GuideTask", None, "GuideTask.xlsx", "id", "map", None, "引导任务/小目标", None, None],
            [None, "TbStaff", "Staff", None, "Staff.xlsx", "id", "map", None, "员工属性", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def create_guide_task_xlsx():
    path = DATAS / "GuideTask.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # param / unlock 为 (list),int；Excel 单元格内用逗号分隔，空表示 []
    # param 按 taskType：Buy*=[targetCount]；Hire*=[targetCount, staffId]
    # unlock：前置任务 Id 列表；空=无前置（招聘类另由 taskType 暗示 Recruit 阶段）
    rows = [
        ["##var", "id", "desc", "taskType", "param", "unlock"],
        ["##type", "int", "string", "GuideTaskType", "(list#sep=,),int", "(list#sep=,),int"],
        ["##group", "c", "c", "c", "c", "c"],
        ["##", "任务Id", "描述文字", "任务类型", "关联参数列表", "解锁前置任务Id列表"],
        [None, 1, "购买基础设施", "BuyBasicEquipment", "1", ""],
        [None, 2, "购买4张桌子", "BuyTables", "4", "1"],
        [None, 3, "购买厨房设施", "BuyKitchenEquipment", "2", "2"],
        [None, 4, "招聘掌柜", "HireShopkeeper", "1,1", "3"],
        [None, 5, "招聘厨师", "HireChef", "1,4", "3"],
        [None, 6, "招聘小二", "HireWaiter", "1,5", "3"],
    ]
    write_rows(ws, rows)
    wb.save(path)
    print(f"created {path}")


def create_staff_xlsx():
    path = DATAS / "Staff.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    rows = [
        [
            "##var",
            "id",
            "name",
            "position",
            "moveSpeed",
            "canOrder",
            "canServe",
            "canCheckout",
            "serviceAttitude",
            "personality",
            "quality",
            "salary",
            "visual",
            "remark",
        ],
        [
            "##type",
            "int",
            "string",
            "StaffPosition",
            "float",
            "bool",
            "bool",
            "bool",
            "int",
            "int",
            "StaffQuality",
            "int",
            "string",
            "string",
        ],
        ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
        [
            "##",
            "员工Id",
            "姓名",
            "职位",
            "移动速度",
            "会点单",
            "会上菜",
            "会收账",
            "服务态度",
            "性格指数",
            "品质",
            "工资",
            "任务形象",
            "备注",
        ],
        # id 与现有引导招聘常量对齐：掌柜1 / 厨师4 / 小二5
        [None, 1, "阿福", "Shopkeeper", 2.2, True, False, True, 80, 60, "Good", 120, "WaiterF1", "默认掌柜，前台接待与收账"],
        [None, 4, "老李", "Chef", 1.8, False, False, False, 70, 55, "Good", 150, "Chef3", "默认厨师，后厨做菜"],
        [None, 5, "小荷", "Waiter", 2.5, True, True, True, 75, 65, "Common", 100, "WaiterF1_1", "默认小二，点单/上菜/收账"],
    ]
    write_rows(ws, rows)
    try:
        wb.save(path)
    except PermissionError:
        print(f"skip {path} (locked)")
        return
    print(f"created {path}")


def main():
    DATAS.mkdir(parents=True, exist_ok=True)
    ensure_enums()
    ensure_beans()
    ensure_tables()
    create_guide_task_xlsx()
    create_staff_xlsx()


if __name__ == "__main__":
    main()
