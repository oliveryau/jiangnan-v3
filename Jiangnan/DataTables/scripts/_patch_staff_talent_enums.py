# -*- coding: utf-8 -*-
"""Append StaffPosition.Management and StaffTalentType 24-38 to __enums__.xlsx."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
ENUMS_PATH = ROOT / "DataTables" / "Datas" / "__enums__.xlsx"

NEW_STAFF_TALENT_TYPES = [
    ("Management", "管理", 24),
    ("WaiterAllServiceSpeedBonusPercent", "全员小二全服务加速", 25),
    ("ChefAllCookingSpeedBonusPercent", "全员厨师烹饪加速", 26),
    ("TeamOccupancyEfficiencyBonus", "团队高客座效率", 27),
    ("TeamCleanSpeedBonusPercent", "团队清扫加速", 28),
    ("TeamServeSpeedBonusPercent", "团队上菜加速", 29),
    ("TeamCheckoutSpeedBonusPercent", "团队收账加速", 30),
    ("TeamOrderSpeedBonusPercent", "团队点单加速", 31),
    ("RecruitmentCostReductionPercent", "招聘费用降低", 32),
    ("DailyWageReductionPercent", "日薪支出降低", 33),
    ("VipAttractionBonusPercent", "贵客吸引力", 34),
    ("TavernAllWorkSpeedBonusPercent", "全店工作效率", 35),
    ("ChefPrepSpeedBonusPercent", "厨师备菜加速", 36),
    ("ExtraDishChancePercent", "额外出菜概率", 37),
    ("IngredientCostReductionPercent", "食材消耗降低", 38),
]


def existing_enum_names(ws) -> set[str]:
    names: set[str] = set()
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, 8).value
        if isinstance(name, str) and name:
            names.add(name)
    return names


def append_staff_position_management(ws, names: set[str]) -> None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "StaffPosition" and ws.cell(row, 8).value == "Management":
            print("StaffPosition.Management already present")
            return

    insert_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 8).value == "Waiter" and ws.cell(row, 10).value == 3:
            insert_row = row + 1
            break

    if insert_row is None:
        raise RuntimeError("StaffPosition.Waiter row not found")

    ws.insert_rows(insert_row)
    ws.cell(insert_row, 2).value = None
    ws.cell(insert_row, 8).value = "Management"
    ws.cell(insert_row, 9).value = "管理"
    ws.cell(insert_row, 10).value = 4
    print(f"inserted StaffPosition.Management at row {insert_row}")


def append_staff_talent_types(ws, names: set[str]) -> None:
    added = 0
    for name, alias, value in NEW_STAFF_TALENT_TYPES:
        if name in names:
            continue
        row = ws.max_row + 1
        ws.cell(row, 8).value = name
        ws.cell(row, 9).value = alias
        ws.cell(row, 10).value = value
        names.add(name)
        added += 1
    print(f"appended {added} StaffTalentType items")


def main() -> None:
    wb = load_workbook(ENUMS_PATH)
    ws = wb.active
    names = existing_enum_names(ws)
    append_staff_position_management(ws, names)
    append_staff_talent_types(ws, names)
    wb.save(ENUMS_PATH)
    print(f"patched {ENUMS_PATH}")


if __name__ == "__main__":
    main()
