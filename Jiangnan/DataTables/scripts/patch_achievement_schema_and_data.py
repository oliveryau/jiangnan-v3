# -*- coding: utf-8 -*-
"""Patch Achievement Luban schema (enums/beans) and normalize Achievement.xlsx data."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[2]
DATAS = ROOT / "DataTables" / "Datas"
DOCS = ROOT / "Docs"

ACHIEVEMENT_TYPES = [
    ("ServeCustomers", "招待客人", 1, "已实现：tavern.totalServedCustomers"),
    ("EarnIncome", "累计赚钱", 2, "已实现：tavern.totalIncome（含柜台随机奖）"),
    ("CookDishes", "做出菜品", 3, "已实现：tavern.totalCookedDishes（偷菜不计）"),
    ("OpenBusiness", "开业次数", 4, "已实现：gameplay.businessOpenCount"),
    ("HireStaff", "同时雇佣员工", 5, "ownedStaff 非临时在职数"),
    ("PromoteManager", "晋升管理员", 6, "未实现：无晋升系统"),
    ("CollectTalent", "收集员工天赋", 7, "achievementStats.unlockedStaffTalentIds（接口预留）"),
    ("ExpandTavern", "扩张桌位", 8, "GetUnlockedTableCount()"),
    ("UnlockDish", "解锁菜品", 9, "未实现：无解锁计数"),
    ("UnlockDrink", "解锁酒水", 10, "未实现：无解锁计数"),
    ("ServeVip", "招待贵宾", 11, "achievementStats.totalVipCheckout"),
    ("ServeVipSuccess", "贵客成功招待", 29, "achievementStats.totalVipSuccessfulServe"),
    ("VipSingleSpendReached", "贵宾单次消费", 12, "achievementStats.peakVipSingleTableIncome"),
    ("VipConcurrentCount", "同时招待贵宾", 13, "achievementStats.peakVipConcurrentCount"),
    ("VipWalkout", "贵宾离场", 14, "achievementStats.totalVipNegativeWalkout"),
    ("PerfectBusinessDay", "完美营业日", 15, "四段等待均 < 5s"),
    ("ServeCustomersOneDay", "单日招待客人", 16, "achievementStats.peakSessionServedCustomers"),
    ("SaveImpatientCustomer", "挽留急客", 17, "未实现：param[0]=满意度阈值 param[1]=人数"),
    ("QueueLengthReached", "排队人数峰值", 18, "achievementStats.peakQueueLength"),
    ("PendingServeDishes", "待上菜峰值", 19, "achievementStats.peakPendingServeDishes"),
    ("PendingCheckoutTables", "待结账峰值", 20, "achievementStats.peakPendingCheckoutTables"),
    ("DirtyTablePeak", "脏桌峰值", 21, "achievementStats.peakDirtyTables"),
    ("SlowServeWalkout", "上菜慢离场", 22, "achievementStats.totalSlowServeWalkout"),
    ("LongWaitWalkout", "等位久离场", 23, "achievementStats.totalLongWaitWalkout"),
    ("ManualServiceActions", "手动服务次数", 24, "achievementStats.totalManualServiceActions"),
    ("AutoServiceDay", "全自动营业日", 25, "achievementStats.autoServiceDayCount"),
    ("HighPriceWalkout", "高价离场", 26, "未实现：无价格敏感离场统计"),
    ("NegativeProfitDay", "负利润营业日", 27, "achievementStats.negativeProfitDayCount"),
    ("CompleteAchievements", "完成成就数量", 28, "可部分实现：统计已达成未领+已领"),
]

ACHIEVEMENT_CATEGORIES = [
    ("BusinessMilestone", "经营里程碑", 1),
    ("StaffGrowth", "员工成长", 2),
    ("Expansion", "扩张经营", 3),
    ("ContentUnlock", "内容解锁", 4),
    ("VipGameplay", "贵宾玩法", 5),
    ("Challenge", "挑战成就", 6),
]

# id, name, desc, type, param, rewardCoin, category, remark
ACHIEVEMENT_ROWS = [
    (1, "新店开张", "累计开业 1 次", "OpenBusiness", "1", 50, "BusinessMilestone", "经营里程碑"),
    (2, "生意兴隆", "累计开业 5 次", "OpenBusiness", "5", 150, "BusinessMilestone", "经营里程碑"),
    (3, "宾客如云", "累计开业 10 次", "OpenBusiness", "10", 500, "BusinessMilestone", "经营里程碑"),
    (4, "百年老店", "累计开业 30 次", "OpenBusiness", "30", 2000, "BusinessMilestone", "经营里程碑"),
    (5, "初出茅庐", "累计招待 10 位客人", "ServeCustomers", "10", 50, "BusinessMilestone", "经营里程碑"),
    (6, "宾客盈门", "累计招待 50 位客人", "ServeCustomers", "50", 200, "BusinessMilestone", "经营里程碑"),
    (7, "高朋满座", "累计招待 200 位客人", "ServeCustomers", "200", 500, "BusinessMilestone", "经营里程碑"),
    (8, "门庭若市", "累计招待 500 位客人", "ServeCustomers", "500", 2000, "BusinessMilestone", "经营里程碑"),
    (9, "小有积蓄", "生涯累计赚取 200 铜钱", "EarnIncome", "200", 100, "BusinessMilestone", "经营里程碑"),
    (10, "日进斗金", "生涯累计赚取 6000 铜钱", "EarnIncome", "6000", 300, "BusinessMilestone", "经营里程碑"),
    (11, "富甲一方", "生涯累计赚取 50000 铜钱", "EarnIncome", "50000", 1000, "BusinessMilestone", "经营里程碑"),
    (12, "富可敌国", "生涯累计赚取 100000 铜钱", "EarnIncome", "100000", 5000, "BusinessMilestone", "经营里程碑"),
    (13, "初试厨技", "累计做出 20 道菜", "CookDishes", "20", 60, "BusinessMilestone", "经营里程碑"),
    (14, "灶火连天", "累计做出 100 道菜", "CookDishes", "100", 250, "BusinessMilestone", "经营里程碑"),
    (15, "珍馐千席", "累计做出 500 道菜", "CookDishes", "500", 800, "BusinessMilestone", "经营里程碑"),
    (16, "御膳名厨", "累计做出 2000 道菜", "CookDishes", "2000", 1500, "BusinessMilestone", "经营里程碑"),
    (17, "广纳贤才", "同时雇佣 3 位员工", "HireStaff", "3", 50, "StaffGrowth", "员工成长"),
    (18, "人才济济", "同时雇佣 6 位员工", "HireStaff", "6", 600, "StaffGrowth", "员工成长"),
    (19, "卧虎藏龙", "累计解锁 5 种不同员工天赋", "CollectTalent", "5", 500, "StaffGrowth", "员工成长"),
    (20, "扩店经营", "解锁 1 张桌子", "ExpandTavern", "1", 100, "Expansion", "扩张经营"),
    (21, "连锁名酒楼", "累计放置 4 张桌子", "ExpandTavern", "4", 500, "Expansion", "扩张经营"),
    (22, "贵客临门", "累计招待 1 位贵宾", "ServeVip", "1", 100, "VipGameplay", "贵宾玩法"),
    (23, "座上宾", "累计成功招待 10 位贵宾", "ServeVipSuccess", "10", 250, "VipGameplay", "贵宾玩法"),
    (24, "一掷千金", "单桌贵宾消费达到 500 铜钱", "VipSingleSpendReached", "500", 300, "VipGameplay", "贵宾玩法"),
    (25, "贵宾云集", "同时接待 3 位贵宾", "VipConcurrentCount", "3", 500, "VipGameplay", "贵宾玩法"),
    (26, "贵客不候", "累计 1 位贵宾负向离场", "VipWalkout", "1", 50, "VipGameplay", "贵宾玩法"),
    (27, "完美的一天", "完成 1 次营业日，每位顾客四段等待均不超过 5 秒", "PerfectBusinessDay", "1", 300, "Challenge", "挑战成就"),
    (28, "翻台如流水", "单日成功招待 30 位客人", "ServeCustomersOneDay", "30", 200, "Challenge", "挑战成就"),
    (29, "排队长龙", "门口排队人数同时达到 10", "QueueLengthReached", "10", 150, "Challenge", "挑战成就"),
    (30, "热的会发慌", "同时有 5 份已出菜品等待上菜", "PendingServeDishes", "5", 50, "Challenge", "挑战成就"),
    (31, "没钱还想走？", "同时有 5 桌客人等待结账", "PendingCheckoutTables", "5", 50, "Challenge", "挑战成就"),
    (32, "桌子必须还我", "同时有 5 桌没清理", "DirtyTablePeak", "5", 50, "Challenge", "挑战成就"),
    (33, "磨叽大师", "累计 5 位顾客因上菜过慢而离场", "SlowServeWalkout", "5", 50, "Challenge", "挑战成就"),
    (34, "这店竟然这么火？", "累计 3 位顾客因等待排队过久而离场", "LongWaitWalkout", "3", 50, "Challenge", "挑战成就"),
    (35, "全能店老板", "累计手动完成 30 次点单+上菜+收账", "ManualServiceActions", "30", 50, "Challenge", "挑战成就"),
    (36, "甩手掌柜", "完成 1 次全程无手动操作的营业日", "AutoServiceDay", "1", 50, "Challenge", "挑战成就"),
    (37, "赔本赚吆喝", "完成 1 次净利润为负的营业日", "NegativeProfitDay", "1", 50, "Challenge", "挑战成就"),
    (38, "模拟经营高手", "完成 9 项成就", "CompleteAchievements", "9", 50, "Challenge", "挑战成就"),
    (39, "模拟经营大师", "完成 19 项成就", "CompleteAchievements", "19", 50, "Challenge", "挑战成就"),
    (40, "这游戏是你做的？", "完成 37 项成就", "CompleteAchievements", "37", 50, "Challenge", "挑战成就"),
]

IMPLEMENTED_TYPES = {
    "ServeCustomers", "EarnIncome", "CookDishes", "OpenBusiness", "CompleteAchievements",
    "ExpandTavern", "HireStaff", "CollectTalent", "ServeVip", "ServeVipSuccess",
    "VipSingleSpendReached", "VipConcurrentCount", "VipWalkout", "PerfectBusinessDay",
    "ServeCustomersOneDay", "QueueLengthReached", "PendingServeDishes", "PendingCheckoutTables",
    "DirtyTablePeak", "SlowServeWalkout", "LongWaitWalkout", "ManualServiceActions",
    "AutoServiceDay", "NegativeProfitDay",
}


def parse_int_list(text: str) -> list[int]:
    text = (text or "").strip()
    if not text:
        return []
    return [int(p.strip()) for p in text.split(",") if p.strip()]


def patch_enums(ws: Worksheet) -> None:
    # Remove existing AchievementType / AchievementCategory blocks
    rows_to_clear = []
    active_enum = None
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, 2).value
        item = ws.cell(row, 8).value
        if name in {"AchievementType", "AchievementCategory"}:
            active_enum = name
            rows_to_clear.append(row)
            continue
        if active_enum and item:
            rows_to_clear.append(row)
            continue
        if active_enum and not item:
            active_enum = None

    for row in sorted(rows_to_clear, reverse=True):
        ws.delete_rows(row, 1)

    start = ws.max_row + 1
    for index, (name, alias, value, comment) in enumerate(ACHIEVEMENT_TYPES):
        row = start + index
        ws.cell(row, 2).value = "AchievementType" if index == 0 else None
        if index == 0:
            ws.cell(row, 4).value = True
            ws.cell(row, 6).value = "成就任务类型"
        ws.cell(row, 8).value = name
        ws.cell(row, 9).value = alias
        ws.cell(row, 10).value = value
        ws.cell(row, 11).value = comment

    start = ws.max_row + 1
    for index, (name, alias, value) in enumerate(ACHIEVEMENT_CATEGORIES):
        row = start + index
        ws.cell(row, 2).value = "AchievementCategory" if index == 0 else None
        if index == 0:
            ws.cell(row, 4).value = True
            ws.cell(row, 6).value = "成就分类"
        ws.cell(row, 8).value = name
        ws.cell(row, 9).value = alias
        ws.cell(row, 10).value = value


def patch_beans(ws: Worksheet) -> None:
    fields = [
        ("id", "int", "成就Id"),
        ("name", "string", "名称"),
        ("desc", "string", "描述"),
        ("achievementType", "AchievementType", "成就类型"),
        ("param", "(list#sep=,),int", "目标参数(按类型解释，通常取param[0])"),
        ("rewardCoin", "int", "领取铜钱奖励"),
        ("sortOrder", "int", "列表排序"),
        ("category", "AchievementCategory", "成就分类"),
        ("remark", "string", "备注/策划说明"),
        ("icon", "string", "图鉴图标Resources路径"),
        ("frame", "string", "图鉴边框Resources路径"),
    ]

    start_row = None
    next_bean_row = None
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if name == "Achievement":
            start_row = row
            continue
        if start_row is not None and name not in (None, ""):
            next_bean_row = row
            break

    if start_row is None:
        start_row = ws.max_row + 1
        next_bean_row = start_row

    delete_count = (next_bean_row - start_row) if next_bean_row is not None else 0
    if delete_count > 0:
        ws.delete_rows(start_row, delete_count)

    for index, (fname, ftype, comment) in enumerate(fields):
        row = start_row + index
        ws.cell(row, 2).value = "Achievement" if index == 0 else None
        ws.cell(row, 7).value = "经营成就任务" if index == 0 else None
        ws.cell(row, 10).value = fname
        ws.cell(row, 12).value = ftype
        ws.cell(row, 14).value = comment

    # Ensure beans after Achievement remain intact.
    dish_fields = [
        ("id", "int", "菜品Id"),
        ("name", "string", "菜名"),
        ("materials", "(list#sep=,),string", "材料标签"),
        ("flavor", "(list#sep=,),string", "口味标签"),
        ("icon", "string", "图标Resources路径"),
        ("summary", "string", "简单描述"),
    ]
    hint_fields = [
        ("id", "int", "模板Id"),
        ("text", "string", "提示语模板"),
    ]

    def ensure_bean(bean_name: str, comment: str, bean_fields: list[tuple[str, str, str]]) -> None:
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 2).value == bean_name:
                return

        start = ws.max_row + 1
        for index, (fname, ftype, field_comment) in enumerate(bean_fields):
            row = start + index
            ws.cell(row, 2).value = bean_name if index == 0 else None
            ws.cell(row, 7).value = comment if index == 0 else None
            ws.cell(row, 10).value = fname
            ws.cell(row, 12).value = ftype
            ws.cell(row, 14).value = field_comment

    ensure_bean("Dish", "菜品配置", dish_fields)
    ensure_bean("VipGuestDemandHint", "贵客猜菜提示语", hint_fields)


def write_achievement_xlsx() -> None:
    path = DATAS / "Achievement.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        ("##var", "id", "name", "desc", "achievementType", "param", "rewardCoin", "sortOrder", "category", "remark"),
        ("##type", "int", "string", "string", "AchievementType", "(list#sep=,),int", "int", "int", "AchievementCategory", "string"),
        ("##group",) + ("c",) * 9,
        ("##", "成就Id", "名称", "描述", "类型", "参数", "铜钱奖励", "排序", "分类", "备注"),
    ]
    for r, row in enumerate(headers, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val

    for index, row in enumerate(ACHIEVEMENT_ROWS):
        r = 5 + index
        achievement_id, name, desc, atype, param, reward, category, remark = row
        ws.cell(r, 2).value = achievement_id
        ws.cell(r, 3).value = name
        ws.cell(r, 4).value = desc
        ws.cell(r, 5).value = atype
        ws.cell(r, 6).value = param
        ws.cell(r, 7).value = reward
        ws.cell(r, 8).value = achievement_id * 10
        ws.cell(r, 9).value = category
        ws.cell(r, 10).value = remark

    try:
        wb.save(path)
        print(f"saved {path}")
    except PermissionError:
        pending = DATAS / "Achievement.pending.xlsx"
        wb.save(pending)
        print(f"WARNING: {path} is locked (close Excel). Wrote {pending} — close file and rerun, or copy over.")


def write_audit_doc() -> None:
    lines = [
        "# 成就系统盘点（自动生成）",
        "",
        "数据源：`DataTables/Datas/Achievement.xlsx`",
        "",
        "## 已实现统计的类型",
        "",
        "| 类型 | 计数来源 | 备注 |",
        "|------|----------|------|",
        "| ServeCustomers | `tavern.totalServedCustomers` | 每次结账 +1 |",
        "| EarnIncome | `tavern.totalIncome` | 含柜台随机奖 |",
        "| CookDishes | `tavern.totalCookedDishes` | 小二偷菜不计 |",
        "| OpenBusiness | `gameplay.businessOpenCount` | 每次开业 +1 |",
        "| HireStaff | `gameplay.ownedStaff.Count` | 已接入读取，招聘时需 Notify |",
        "| ExpandTavern | `GetUnlockedTableCount()` | 已接入读取 |",
        "| CompleteAchievements | 已达成成就数 | 不含 CompleteAchievements 自身 |",
        "",
        "## 未实现统计的类型（当前进度恒为 0）",
        "",
    ]
    for name, alias, _value, comment in ACHIEVEMENT_TYPES:
        if name in IMPLEMENTED_TYPES:
            continue
        lines.append(f"- **{name}**（{alias}）：{comment}")

    lines.extend(["", "## 配表成就清单", ""])
    lines.extend([
        "| id | 名称 | 类型 | 目标 | 分类 | 可达成 |",
        "|----|------|------|------|------|--------|",
    ])
    for row in ACHIEVEMENT_ROWS:
        achievement_id, name, _desc, atype, param, _reward, category, _remark = row
        ok = "是" if atype in IMPLEMENTED_TYPES else "否"
        lines.append(f"| {achievement_id} | {name} | {atype} | {param} | {category} | {ok} |")

    lines.extend([
        "",
        "## Bean 字段",
        "",
        "| 字段 | 类型 | 说明 |",
        "|------|------|------|",
        "| id | int | 成就 Id |",
        "| name | string | 显示名 |",
        "| desc | string | 描述 |",
        "| achievementType | AchievementType | 条件类型 |",
        "| param | list,int | 目标参数 |",
        "| rewardCoin | int | 领取奖励 |",
        "| sortOrder | int | 列表排序 |",
        "| category | AchievementCategory | 分类（原 remark 分类结构化） |",
        "| remark | string | 策划备注 |",
        "",
        "## 已知问题",
        "",
        "- 原 Excel 存在 id 13–16 重复，已重编号为 17–45。",
        "- `sortOrder` 原为空，已按 id×10 填充。",
        "- 成就入口 UI 尚未接入游戏内按钮。",
    ])

    out = DOCS / "成就系统盘点.md"
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"saved {out}")


def main() -> None:
    enums_path = DATAS / "__enums__.xlsx"
    beans_path = DATAS / "__beans__.xlsx"

    wb_enums = load_workbook(enums_path)
    patch_enums(wb_enums.active)
    wb_enums.save(enums_path)
    print(f"patched {enums_path}")

    wb_beans = load_workbook(beans_path)
    patch_beans(wb_beans.active)
    wb_beans.save(beans_path)
    print(f"patched {beans_path}")

    write_achievement_xlsx()
    # write_audit_doc() — 手工维护 Docs/成就系统盘点.md，避免覆盖


if __name__ == "__main__":
    main()
