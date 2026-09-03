#!/usr/bin/env python3
"""Expand StaffLevel (waiter 8-tier) + TavernTech (3 branches) and register schema."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATAS = ROOT / "Datas"


def write_rows(ws, rows: list[list]):
    ws.delete_rows(1, ws.max_row or 1)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val


def ensure_enums():
    path = DATAS / "__enums__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "flags", "unique", "group", "comment", "tags", "*items", None, None, None, None],
            ["##var", None, None, None, None, None, None, "name", "alias", "value", "comment", "tags"],
            ["##", "全名(包含模块名可空)", "是否为位标枚举", "枚举项是否唯一", None, None, None, "枚举名", "别名", "值", "备注", None],
            [None, "GuideTaskType", None, True, None, "引导任务类型", None, "BuyBasicEquipment", "购买基础设施", 1, None, None],
            [None, None, None, None, None, None, None, "BuyTables", "购买桌子", 2, None, None],
            [None, None, None, None, None, None, None, "BuyKitchenEquipment", "购买厨房设施", 3, None, None],
            [None, None, None, None, None, None, None, "HireShopkeeper", "招聘掌柜", 4, None, None],
            [None, None, None, None, None, None, None, "HireChef", "招聘厨师", 5, None, None],
            [None, None, None, None, None, None, None, "HireWaiter", "招聘小二", 6, None, None],
            [None, None, None, None, None, None, None, "Custom", "自定义", 99, None, None],
            [None, "StaffPosition", None, True, None, "员工职位", None, "Shopkeeper", "掌柜", 1, None, None],
            [None, None, None, None, None, None, None, "Chef", "厨师", 2, None, None],
            [None, None, None, None, None, None, None, "Waiter", "小二", 3, None, None],
            [None, "StaffQuality", None, True, None, "员工品质", None, "Common", "普通", 1, None, None],
            [None, None, None, None, None, None, None, "Good", "优良", 2, None, None],
            [None, None, None, None, None, None, None, "Rare", "稀有", 3, None, None],
            [None, None, None, None, None, None, None, "Epic", "史诗", 4, None, None],
            [None, "FacilityType", None, True, None, "设施类型", None, "Table", "桌子", 1, None, None],
            [None, None, None, None, None, None, None, "Counter", "掌柜桌", 2, None, None],
            [None, None, None, None, None, None, None, "Stove", "灶台", 3, None, None],
            [None, None, None, None, None, None, None, "Furnace", "炉子", 4, None, None],
            [None, None, None, None, None, None, None, "WineCabinet", "酒柜", 5, None, None],
            [None, None, None, None, None, None, None, "Cabinet", "柜子", 6, None, None],
            [None, None, None, None, None, None, None, "KitchenTable", "厨房桌", 7, None, None],
            # TavernTechType — 编制 / 客流 / 经营
            [None, "TavernTechType", None, True, None, "酒馆科技效果类型", None, "ExtraWaiterCap", "增加小二上限", 1, None, None],
            [None, None, None, None, None, None, None, "ExtraChefCap", "增加厨师上限", 2, None, None],
            [None, None, None, None, None, None, None, "ExtraShopkeeperCap", "增加掌柜上限", 3, None, None],
            [None, None, None, None, None, None, None, "QueueCap", "增加排队容量", 4, None, None],
            [None, None, None, None, None, None, None, "BusinessHoursBonus", "营业时长加成秒", 5, None, None],
            [None, None, None, None, None, None, None, "CustomerRefreshMul", "刷客间隔千分比", 6, None, None],
            [None, None, None, None, None, None, None, "PriceProfitBonus", "涨价利润额外百分比", 7, None, None],
            [None, None, None, None, None, None, None, "InspireBonus", "鼓舞客流额外百分比", 8, None, None],
            [None, None, None, None, None, None, None, "ResearchSpeedMul", "研究耗时千分比", 9, None, None],
            [None, None, None, None, None, None, None, "TipGlobalBonus", "全店小费百分比", 10, None, None],
            [None, None, None, None, None, None, None, "Custom", "自定义", 99, None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_beans():
    path = DATAS / "__beans__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "parent", "valueType", "sep", "alias", "comment", "group", "tags", "*fields", None, None, None, None, None, None],
            ["##var", None, None, None, None, None, None, None, None, "name", "alias", "type", "group", "comment", "tags", "variants"],
            ["##", "全名(包含模块名可空)", None, None, "分隔符", None, None, None, None, "字段名", "字段别名", "类型", "分组", "备注", None, "字段变体"],
            [None, "Config", None, None, None, None, None, None, None, "configName", None, "string", None, "配置名称", None, None],
            [None, None, None, None, None, None, None, None, None, "value", None, "int", None, "数值", None, None],
            [None, "GuideTask", None, None, None, None, "引导小目标任务", None, None, "id", None, "int", None, "任务Id", None, None],
            [None, None, None, None, None, None, None, None, None, "desc", None, "string", None, "描述文字", None, None],
            [None, None, None, None, None, None, None, None, None, "taskType", None, "GuideTaskType", None, "任务类型", None, None],
            [None, None, None, None, None, None, None, None, None, "param", None, "(list#sep=,),int", None, "关联参数列表(按taskType解释)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "解锁前置任务Id列表", None, None],
            [None, "Staff", None, None, None, None, "员工属性", None, None, "id", None, "int", None, "员工Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "姓名", None, None],
            [None, None, None, None, None, None, None, None, None, "position", None, "StaffPosition", None, "职位", None, None],
            [None, None, None, None, None, None, None, None, None, "moveSpeed", None, "float", None, "移动速度", None, None],
            [None, None, None, None, None, None, None, None, None, "canOrder", None, "bool", None, "会点单", None, None],
            [None, None, None, None, None, None, None, None, None, "canServe", None, "bool", None, "会上菜", None, None],
            [None, None, None, None, None, None, None, None, None, "canCheckout", None, "bool", None, "会收账", None, None],
            [None, None, None, None, None, None, None, None, None, "serviceAttitude", None, "int", None, "服务态度指数", None, None],
            [None, None, None, None, None, None, None, None, None, "personality", None, "int", None, "性格指数", None, None],
            [None, None, None, None, None, None, None, None, None, "quality", None, "StaffQuality", None, "品质", None, None],
            [None, None, None, None, None, None, None, None, None, "salary", None, "int", None, "工资", None, None],
            [None, None, None, None, None, None, None, None, None, "visual", None, "string", None, "任务形象/预制体Key", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
            [None, "Facility", None, None, None, None, "酒楼设施", None, None, "id", None, "int", None, "设施Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "显示名", None, None],
            [None, None, None, None, None, None, None, None, None, "facilityType", None, "FacilityType", None, "设施类型", None, None],
            [None, None, None, None, None, None, None, None, None, "bindNode", None, "string", None, "场景绑定节点名", None, None],
            [None, None, None, None, None, None, None, None, None, "guideKey", None, "string", None, "引导逻辑Key", None, None],
            [None, None, None, None, None, None, None, None, None, "equipmentId", None, "int", None, "关联SO_EquipmentId(无则0)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "解锁前置设施Id列表", None, None],
            [None, None, None, None, None, None, None, None, None, "unlockIncome", None, "int", None, "解锁所需累计营收(0忽略)", None, None],
            [None, None, None, None, None, None, None, None, None, "cost", None, "int", None, "解锁花费", None, None],
            [None, None, None, None, None, None, None, None, None, "upgradeGroup", None, "(list#sep=,),int", None, "升级用设备Id组(SO_Equipment)", None, None],
            [None, None, None, None, None, None, None, None, None, "sortOrder", None, "int", None, "同类型排序", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
            # StaffLevel — 含动作耗时倍率 / 小费 / 抗压
            [None, "StaffLevel", None, None, None, None, "员工等级成长", None, None, "id", None, "int", None, "等级配置Id", None, None],
            [None, None, None, None, None, None, None, None, None, "position", None, "StaffPosition", None, "适用职位", None, None],
            [None, None, None, None, None, None, None, None, None, "level", None, "int", None, "等级", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "等级显示名", None, None],
            [None, None, None, None, None, None, None, None, None, "moveSpeedMul", None, "float", None, "移速倍率(相对Staff基值)", None, None],
            [None, None, None, None, None, None, None, None, None, "cookSpeedMul", None, "float", None, "做菜速度倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "orderTimeMul", None, "float", None, "点单耗时倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "serveTimeMul", None, "float", None, "上菜耗时倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "checkoutTimeMul", None, "float", None, "收账耗时倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "cleanTimeMul", None, "float", None, "清扫耗时倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "tipBonusPercent", None, "int", None, "小费加成百分比", None, None],
            [None, None, None, None, None, None, None, None, None, "staminaDrainMul", None, "float", None, "体力消耗倍率", None, None],
            [None, None, None, None, None, None, None, None, None, "canOrder", None, "bool", None, "本级开放点单", None, None],
            [None, None, None, None, None, None, None, None, None, "canServe", None, "bool", None, "本级开放上菜", None, None],
            [None, None, None, None, None, None, None, None, None, "canCheckout", None, "bool", None, "本级开放收账", None, None],
            [None, None, None, None, None, None, None, None, None, "serviceAttitudeBonus", None, "int", None, "服务态度加成", None, None],
            [None, None, None, None, None, None, None, None, None, "personalityBonus", None, "int", None, "性格加成", None, None],
            [None, None, None, None, None, None, None, None, None, "upgradeCost", None, "int", None, "升到本级花费", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
            # TavernTech
            [None, "TavernTech", None, None, None, None, "酒馆科技树", None, None, "id", None, "int", None, "科技Id", None, None],
            [None, None, None, None, None, None, None, None, None, "name", None, "string", None, "科技名", None, None],
            [None, None, None, None, None, None, None, None, None, "desc", None, "string", None, "描述", None, None],
            [None, None, None, None, None, None, None, None, None, "techType", None, "TavernTechType", None, "效果类型", None, None],
            [None, None, None, None, None, None, None, None, None, "param", None, "(list#sep=,),int", None, "效果参数(按techType解释)", None, None],
            [None, None, None, None, None, None, None, None, None, "unlock", None, "(list#sep=,),int", None, "前置科技Id列表", None, None],
            [None, None, None, None, None, None, None, None, None, "cost", None, "int", None, "研究花费", None, None],
            [None, None, None, None, None, None, None, None, None, "researchServeCustomers", None, "int", None, "研究需招待顾客(结算)数量", None, None],
            [None, None, None, None, None, None, None, None, None, "sortOrder", None, "int", None, "排序", None, None],
            [None, None, None, None, None, None, None, None, None, "remark", None, "string", None, "备注", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def ensure_tables():
    path = DATAS / "__tables__.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    write_rows(
        ws,
        [
            ["##var", "full_name", "value_type", "read_schema_from_file", "input", "index", "mode", "group", "comment", "tags", "output"],
            ["##", "全名(包含模块名可空)", "记录类型", "从excel读取定义", "文件列表", "主键字段", "模式", "分组", "备注", None, "输出文件名"],
            ["##", None, None, None, None, None, None, None, None, None, None],
            [None, "TbConfig", "Config", None, "Config.xlsx", None, "list", None, "全局KV配置", None, None],
            [None, "TbGuideTask", "GuideTask", None, "GuideTask.xlsx", "id", "map", None, "引导任务/小目标", None, None],
            [None, "TbStaff", "Staff", None, "Staff.xlsx", "id", "map", None, "员工属性", None, None],
            [None, "TbFacility", "Facility", None, "Facility.xlsx", "id", "map", None, "酒楼设施", None, None],
            [None, "TbStaffLevel", "StaffLevel", None, "StaffLevel.xlsx", "id", "map", None, "员工等级成长", None, None],
            [None, "TbTavernTech", "TavernTech", None, "TavernTech.xlsx", "id", "map", None, "酒馆科技树", None, None],
        ],
    )
    wb.save(path)
    print(f"updated {path}")


def create_staff_level_xlsx():
    path = DATAS / "StaffLevel.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header = [
        [
            "##var",
            "id",
            "position",
            "level",
            "name",
            "moveSpeedMul",
            "cookSpeedMul",
            "orderTimeMul",
            "serveTimeMul",
            "checkoutTimeMul",
            "cleanTimeMul",
            "tipBonusPercent",
            "staminaDrainMul",
            "canOrder",
            "canServe",
            "canCheckout",
            "serviceAttitudeBonus",
            "personalityBonus",
            "upgradeCost",
            "remark",
        ],
        [
            "##type",
            "int",
            "StaffPosition",
            "int",
            "string",
            "float",
            "float",
            "float",
            "float",
            "float",
            "float",
            "int",
            "float",
            "bool",
            "bool",
            "bool",
            "int",
            "int",
            "int",
            "string",
        ],
        ["##group"] + ["c"] * 19,
        [
            "##",
            "配置Id",
            "职位",
            "等级",
            "显示名",
            "移速",
            "做菜",
            "点单耗时",
            "上菜耗时",
            "收账耗时",
            "清扫耗时",
            "小费%",
            "体力消耗",
            "点单",
            "上菜",
            "收账",
            "态度+",
            "性格+",
            "升级费",
            "备注",
        ],
    ]
    # Waiter 8 tiers
    waiter = [
        [None, 101, "Waiter", 1, "见习小二", 1.00, 1.0, 1.00, 1.00, 1.00, 1.00, 0, 1.00, True, False, False, 0, 0, 0, "入职仅点单"],
        [None, 102, "Waiter", 2, "跑堂小二", 1.05, 1.0, 0.95, 1.00, 1.00, 0.95, 0, 1.00, True, True, False, 3, 0, 200, "开放上菜"],
        [None, 103, "Waiter", 3, "熟练小二", 1.10, 1.0, 0.90, 0.95, 1.00, 0.90, 2, 0.95, True, True, True, 6, 2, 400, "开放收账"],
        [None, 104, "Waiter", 4, "利落小二", 1.15, 1.0, 0.85, 0.90, 0.90, 0.85, 4, 0.95, True, True, True, 8, 4, 700, "全流程加速"],
        [None, 105, "Waiter", 5, "老练小二", 1.20, 1.0, 0.80, 0.85, 0.85, 0.80, 6, 0.90, True, True, True, 12, 6, 1100, "中坚"],
        [None, 106, "Waiter", 6, "金牌小二", 1.25, 1.0, 0.75, 0.80, 0.80, 0.75, 8, 0.90, True, True, True, 15, 8, 1600, "小费明显"],
        [None, 107, "Waiter", 7, "领班小二", 1.30, 1.0, 0.70, 0.75, 0.75, 0.70, 12, 0.85, True, True, True, 18, 10, 2200, "高峰抗压"],
        [None, 108, "Waiter", 8, "店中翘楚", 1.35, 1.0, 0.65, 0.70, 0.70, 0.65, 15, 0.80, True, True, True, 22, 12, 3000, "满级"],
    ]
    # Chef / Shopkeeper keep 3 tiers with new columns defaulted
    chef = [
        [None, 201, "Chef", 1, "见习厨师", 1.00, 1.00, 1.0, 1.0, 1.0, 1.0, 0, 1.0, False, False, False, 0, 0, 0, ""],
        [None, 202, "Chef", 2, "熟练厨师", 1.05, 1.15, 1.0, 1.0, 1.0, 1.0, 0, 0.95, False, False, False, 0, 5, 400, ""],
        [None, 203, "Chef", 3, "金牌厨师", 1.10, 1.30, 1.0, 1.0, 1.0, 1.0, 0, 0.90, False, False, False, 5, 10, 1000, ""],
    ]
    shopkeeper = [
        [None, 301, "Shopkeeper", 1, "见习掌柜", 1.00, 1.0, 1.0, 1.0, 1.00, 1.0, 0, 1.0, False, False, True, 0, 0, 0, "默认可收账"],
        [None, 302, "Shopkeeper", 2, "熟练掌柜", 1.05, 1.0, 1.0, 1.0, 0.90, 1.0, 3, 0.95, False, False, True, 5, 5, 500, ""],
        [None, 303, "Shopkeeper", 3, "金牌掌柜", 1.10, 1.0, 1.0, 1.0, 0.80, 1.0, 6, 0.90, False, False, True, 10, 10, 1200, ""],
    ]
    write_rows(ws, header + waiter + chef + shopkeeper)
    wb.save(path)
    print(f"created {path}")


def create_tavern_tech_xlsx():
    path = DATAS / "TavernTech.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header = [
        [
            "##var",
            "id",
            "name",
            "desc",
            "techType",
            "param",
            "unlock",
            "cost",
            "researchServeCustomers",
            "sortOrder",
            "remark",
        ],
        [
            "##type",
            "int",
            "string",
            "string",
            "TavernTechType",
            "(list#sep=,),int",
            "(list#sep=,),int",
            "int",
            "int",
            "int",
            "string",
        ],
        ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
        [
            "##",
            "科技Id",
            "名称",
            "描述",
            "效果类型",
            "参数",
            "前置科技",
            "花费",
            "研究秒数",
            "排序",
            "备注",
        ],
    ]
    # Branch 1 编制 100-104
    # Branch 2 客流 200-204
    # Branch 3 经营 300-304
    rows = header + [
        [None, 100, "扩招小二I", "小二招聘上限+1", "ExtraWaiterCap", "1", "", 500, 30, 100, "编制"],
        [None, 101, "扩招小二II", "小二招聘上限再+1", "ExtraWaiterCap", "1", "100", 1000, 45, 101, "编制"],
        [None, 102, "扩招小二III", "小二招聘上限再+1", "ExtraWaiterCap", "1", "101", 1800, 60, 102, "编制"],
        [None, 103, "扩招厨师I", "厨师招聘上限+1", "ExtraChefCap", "1", "", 800, 40, 103, "编制"],
        [None, 104, "扩招厨师II", "厨师招聘上限再+1", "ExtraChefCap", "1", "103", 1500, 60, 104, "编制"],
        [None, 200, "门庭开阔", "排队容量+5", "QueueCap", "5", "", 600, 35, 200, "客流"],
        [None, 201, "延时打烊I", "营业时长+60秒", "BusinessHoursBonus", "60", "200", 900, 40, 201, "客流"],
        [None, 202, "客似云来I", "刷客间隔×0.9", "CustomerRefreshMul", "900", "200", 1100, 50, 202, "param千分比"],
        [None, 203, "延时打烊II", "营业时长再+90秒", "BusinessHoursBonus", "90", "201", 1600, 55, 203, "客流"],
        [None, 204, "客似云来II", "刷客间隔×0.85", "CustomerRefreshMul", "850", "202", 2000, 70, 204, "param千分比"],
        [None, 300, "精打细算", "全店小费+5%", "TipGlobalBonus", "5", "", 700, 35, 300, "经营"],
        [None, 301, "时来运转", "研究耗时×0.85", "ResearchSpeedMul", "850", "300", 1000, 40, 301, "param千分比"],
        [None, 302, "大胆涨价", "涨价额外利润+25%", "PriceProfitBonus", "25", "300", 1200, 45, 302, "经营"],
        [None, 303, "锣鼓喧天", "鼓舞客流额外+25%", "InspireBonus", "25", "302", 1500, 55, 303, "经营"],
        [None, 304, "财源广进", "全店小费再+10%", "TipGlobalBonus", "10", "301,302", 2200, 75, 304, "需301+302"],
    ]
    write_rows(ws, rows)
    wb.save(path)
    print(f"created {path}")


def main():
    DATAS.mkdir(parents=True, exist_ok=True)
    ensure_enums()
    ensure_beans()
    ensure_tables()
    create_staff_level_xlsx()
    create_tavern_tech_xlsx()


if __name__ == "__main__":
    main()
